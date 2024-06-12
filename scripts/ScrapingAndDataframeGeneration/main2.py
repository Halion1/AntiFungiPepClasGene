from requests_html import HTMLSession
from bs4 import BeautifulSoup
import openpyxl as xl
import lxml.html as lh
import pandas as pd

# READING OUR DATABASE
database = xl.load_workbook('peptidesforfungi2.xlsx')
sheet = database['Sheet1']
len_max=sheet.max_row

for i in range(2, 4):
    # ENTERING TO THE WEBSITE according to the ID
    a=sheet.cell(i,1)
    url="https://dbaasp.org/peptide-card?id="+str(a.value)

    session = HTMLSession()

    r = session.get(url)
    r.html.render()

    soup = BeautifulSoup(r.html.html, "lxml")
    infotable = soup.find('div', class_="container-fluid physico-chemical-property-container m-t-50")
    print(type(infotable))
    if infotable is None :
        print(0)
    else:
        table = infotable.find("tr", {"class": "odd"})
        values = [tag.text for tag in table]
        print(values)

