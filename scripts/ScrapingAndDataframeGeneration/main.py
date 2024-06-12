#import stuff

from requests_html import HTMLSession
from bs4 import BeautifulSoup
import openpyxl as xl
import time

database = xl.load_workbook('peptidesforfungi2.xlsx')
sheet = database['Sheet1']


#obtener columna de ID
cell = sheet.cell(3,1) #fila y columna
length = sheet.max_row

array = []
#length+1

for i in range (2,4):
    a = sheet.cell(i,1)
    url = "https://dbaasp.org/peptide-card?id=" + str(a.value)
    session = HTMLSession()

    r = session.get(url)
    r.html.render()

    soup = BeautifulSoup(r.html.html, "lxml")
    infotable = soup.find('div', class_="container-fluid physico-chemical-property-container m-t-50")
    table = infotable.find("tr", {"class": "odd"})
    values = [tag.text for tag in table]

    print(values)

#print(array)

#usar BeautifulSoup

#soup = BeautifulSoup(page.content, 'htlm.parser')

#url = "https://dbaasp.org/peptide-card?id=" + str(cell.value)

#page = requests.get(url)
#time.sleep(10)
#soup = BeautifulSoup(page.content, 'lxml')

